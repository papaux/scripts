from icalevents.icalevents import events
from datetime import datetime
import openpyxl

#  pip install openpyxl icalevents

def ics_to_excel(ics_file, excel_file):
    today = datetime.today()
    future = datetime(today.year + 2, today.month, today.day)  # limite arbitraire 2 ans

    # Extraire tous les événements (y compris récurrents) entre today et future
    all_events = events(file=ics_file, start=today, end=future)

    # Trier par date
    events_list = sorted([(e.start.date(), e.summary) for e in all_events], key=lambda x: x[0])

    # Créer Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    for col, (_, summary) in enumerate(events_list, start=1):
        ws.cell(row=1, column=col, value=summary)

    for col, (date, _) in enumerate(events_list, start=1):
        ws.cell(row=2, column=col, value=date.strftime("%d.%m.%Y"))

    wb.save(excel_file)
    print(f"Fichier Excel généré : {excel_file}")


# Exemple d’utilisation
ics_to_excel("calendar.ics", "output.xls")

